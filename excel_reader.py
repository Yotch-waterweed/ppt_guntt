"""
Excel → ガントチャートデータ変換モジュール

Excelフォーマット（1行目はヘッダー）:
  A: 大項目    - プロジェクト名（同じ値でグループ化）
  B: 種別      - milestone / period / task
  C: 名前      - アイテム名
  D: 開始日    - YYYY-MM-DD または Excelの日付
  E: 終了日    - YYYY-MM-DD または Excelの日付（milestoneは空可）
"""
import datetime
from typing import List, Dict, Any
from openpyxl import load_workbook


def _to_date_str(value) -> str:
    """セルの値を YYYY-MM-DD 文字列に変換する"""
    if value is None:
        return ""
    if isinstance(value, datetime.datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, datetime.date):
        return value.strftime("%Y-%m-%d")
    s = str(value).strip()
    if not s:
        return ""
    # YYYY/MM/DD 形式もサポート
    return s.replace("/", "-")


def read_excel(filepath: str, sheet_name: str = None) -> List[Dict[str, Any]]:
    """
    Excelファイルを読み込み、GanttChartGenerator が受け取るデータ形式に変換する。
    
    Args:
        filepath: Excelファイルのパス
        sheet_name: シート名（None の場合は最初のシートを使用）
    
    Returns:
        GanttChartGenerator 用のデータリスト
    """
    wb = load_workbook(filepath, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    
    # ヘッダー行をスキップして全行を読み込み
    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # 空行スキップ（大項目が空なら終了）
        if row[0] is None and row[1] is None:
            continue
        
        category = str(row[0]).strip() if row[0] else None
        kind = str(row[1]).strip().lower() if row[1] else ""
        name = str(row[2]).strip() if row[2] else ""
        start_date = _to_date_str(row[3] if len(row) > 3 else None)
        end_date = _to_date_str(row[4] if len(row) > 4 else None)
        
        rows.append({
            "category": category,
            "kind": kind,
            "name": name,
            "start": start_date,
            "end": end_date,
            "row_num": row_idx,
        })
    
    wb.close()
    
    # 大項目でグルーピング（連続する同じ大項目をまとめる）
    data = []
    current_category = None
    current_group = None
    
    for row in rows:
        cat = row["category"]
        
        # 大項目が変わったら新しいグループ
        # 大項目が空の場合は直前のグループに追加
        if cat and cat != current_category:
            if current_group:
                data.append(current_group)
            current_category = cat
            current_group = {
                "name": cat,
                "type": "project",
                "items": [],
                "tasks": [],
            }
        
        if current_group is None:
            continue
        
        kind = row["kind"]
        
        if kind == "milestone":
            date = row["start"] or row["end"]
            if date:
                current_group["items"].append({
                    "name": row["name"],
                    "type": "milestone",
                    "date": date,
                })
        elif kind == "period":
            if row["start"] and row["end"]:
                current_group["items"].append({
                    "name": row["name"],
                    "type": "period",
                    "start": row["start"],
                    "end": row["end"],
                })
        elif kind == "task":
            if row["start"] and row["end"]:
                current_group["tasks"].append({
                    "name": row["name"],
                    "start": row["start"],
                    "end": row["end"],
                })
    
    # 最後のグループを追加
    if current_group:
        data.append(current_group)
    
    return data
